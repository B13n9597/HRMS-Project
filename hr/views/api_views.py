import json
import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from hr.models import (
    Employee, Department, Position, LeaveRequest, LeaveBalance, LeaveType,
    Payroll, PerformanceEvaluation, KPICategory, KPIIndicator, EvaluationScore,
    JobPosting, Application, Applicant, SystemSetting, Role, EmployeeStatus,
    EmployeeHistory, Salary
)

# Helper: check if user is HR/Admin
def is_hr(user):
    # Consider superusers and staff users as HR for API access consistency
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return True
    try:
        employee = Employee.objects.get(user=user)
        if employee.role and employee.role.name.upper() in ('HR', 'ADMIN', 'HR MANAGER', 'DEAN'):
            return True
    except Employee.DoesNotExist:
        pass
    return False


def _employee_payload(employee):
    latest_salary = employee.salaries.order_by('-effective_from').first()
    return {
        'id': employee.id,
        'full_name': employee.get_full_name(),
        'username': employee.user.username if employee.user else '',
        'email': employee.user.email if employee.user else '',
        'role': employee.role.name if employee.role else 'Employee',
        'department': employee.department.name if employee.department else 'General',
        'position': employee.position.title if employee.position else 'Staff',
        'status': employee.status.name if employee.status else 'Active',
        'hire_date': employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else '',
        'phone': employee.phone,
        'salary': str(latest_salary.base_salary) if latest_salary else '',
    }


def _history_payload(history):
    return {
        'id': history.id,
        'employee': history.employee.get_full_name() if history.employee else 'Unknown',
        'event_type': history.get_event_type_display(),
        'old_value': history.old_value,
        'new_value': history.new_value,
        'notes': history.notes,
        'start_date': history.start_date.strftime('%Y-%m-%d') if history.start_date else '',
        'end_date': history.end_date.strftime('%Y-%m-%d') if history.end_date else '',
        'recorded_by': history.recorded_by.get_full_name() if history.recorded_by else '',
    }


# Render the landing home page
def home_view(request):
    try:
        employee_count = Employee.objects.filter(is_deleted=False).count()
        today = timezone.localdate()
        attendance_count = Attendance.objects.filter(date=today, status__in=['Present', 'Late']).count()
        
        evals_count = PerformanceEvaluation.objects.filter(is_deleted=False).count()
        performance_cycle = min(100, int((evals_count / max(1, employee_count)) * 100))
        if performance_cycle == 0:
            performance_cycle = 75
    except Exception:
        employee_count = 0
        attendance_count = 0
        performance_cycle = 75

    context = {
        'employee_count': employee_count,
        'attendance_count': attendance_count,
        'performance_cycle': performance_cycle,
    }
    return render(request, 'hr/home.html', context)

# Render the login page
def login_view(request):
    if request.user.is_authenticated:
        return redirect('/attendance/')
    return render(request, 'hr/login.html')

# Render the public kiosk page
def kiosk_view(request):
    return render(request, 'hr/kiosk.html')

# Public API: List employees for Kiosk dropdown
def api_kiosk_employees(request):
    employees = Employee.objects.filter(is_deleted=False)
    data = []
    for emp in employees:
        data.append({
            'id': emp.id,
            'full_name': emp.get_full_name(),
            'department': emp.department.name if emp.department else 'General',
        })
    return JsonResponse({'success': True, 'employees': data})

# Public API: Register scan from Kiosk
def api_kiosk_scan(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
        
    employee = get_object_or_404(Employee, id=employee_id)
    
    now = timezone.now()
    now_local = timezone.localtime(now)
    today = timezone.localdate()
    
    from hr.models import Attendance, QRScanLog
    attendance, created = Attendance.objects.get_or_create(
        employee=employee,
        date=today,
        defaults={'time_in': now, 'status': 'Present'},
    )
    
    if created:
        attendance.calculate_status()
        attendance.save()
        QRScanLog.objects.create(
            employee=employee, scan_type='time_in', is_successful=True, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'time_in',
                'employee': employee.get_full_name(),
                'time': now_local.strftime('%H:%M:%S'),
                'message': f"Welcome, {employee.first_name}! Clocked in successfully."
            }
        })
    elif attendance.time_out is None:
        attendance.time_out = now
        attendance.save()
        hours = round((now - attendance.time_in).total_seconds() / 3600, 2)
        QRScanLog.objects.create(
            employee=employee, scan_type='time_out', is_successful=True, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'time_out',
                'employee': employee.get_full_name(),
                'time': now_local.strftime('%H:%M:%S'),
                'hours_worked': hours,
                'message': f"Goodbye, {employee.first_name}! You worked {hours} hours."
            }
        })
    else:
        QRScanLog.objects.create(
            employee=employee, scan_type='duplicate', is_successful=False, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'already_complete',
                'employee': employee.get_full_name(),
                'message': 'Attendance already fully recorded for today.'
            }
        })

# Render the single page application
@login_required(login_url='/login/')
@ensure_csrf_cookie
def index_view(request):
    return render(request, 'hr/index.html')

# API: Current User Profile
def api_me(request):
    if not request.user.is_authenticated:
        return JsonResponse({'authenticated': False})
    
    user = request.user
    user_data = {
        'username': user.username,
        'email': user.email,
        'is_superuser': user.is_superuser,
    }
    
    try:
        employee = Employee.objects.get(user=user)
        user_data.update({
            'employee_id': employee.id,
            'first_name': employee.first_name,
            'last_name': employee.last_name,
            'full_name': employee.get_full_name(),
            'phone': employee.phone,
            'address': employee.address,
            'qr_token': str(employee.qr_token),
            'role': employee.role.name if employee.role else 'Employee',
            'department': employee.department.name if employee.department else None,
            'position': employee.position.title if employee.position else None,
            'hire_date': employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else None,
            'is_hr': is_hr(user)
        })
    except Employee.DoesNotExist:
        user_data.update({
            'employee_id': None,
            'full_name': user.get_full_name() or user.username,
            'is_hr': is_hr(user)
        })
        
    return JsonResponse({'authenticated': True, 'user': user_data})

# API: Custom JSON Login
def api_login(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
        
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        # Fetch profile
        return api_me(request)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid username or password'}, status=401)

# API: Custom JSON Logout
def api_logout(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    logout(request)
    return JsonResponse({'success': True})

# API: Employees Directory (HR Only)
@login_required
def api_employees(request):
    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
    employees = Employee.objects.select_related('role', 'department', 'position', 'status').all()
    data = []
    for emp in employees:
        data.append({
            'id': emp.id,
            'employee_id': emp.employee_id,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'full_name': emp.get_full_name(),
            'email': emp.user.email if emp.user else '',
            'username': emp.user.username if emp.user else '',
            'role': emp.role.name if emp.role else 'Employee',
            'department': emp.department.name if emp.department else 'N/A',
            'position': emp.position.title if emp.position else 'N/A',
            'status': emp.status.name if emp.status else 'Active',
            'hire_date': emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '',
            'phone': emp.phone,
            'address': emp.address,
        })
    return JsonResponse({'success': True, 'employees': data})


@login_required
def api_lifecycle(request):
    try:
        current_employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        current_employee = None

    if is_hr(request.user):
        employees = Employee.objects.select_related('user', 'role', 'department', 'position', 'status').all()
        employee_id = request.GET.get('employee_id')
        selected = employees.filter(id=employee_id).first() if employee_id else employees.first()
    else:
        employees = Employee.objects.filter(id=current_employee.id) if current_employee else Employee.objects.none()
        selected = current_employee

    history = EmployeeHistory.objects.none()
    if selected:
        history = EmployeeHistory.objects.filter(employee=selected).select_related('employee', 'recorded_by').order_by('-start_date')

    return JsonResponse({
        'success': True,
        'employees': [_employee_payload(employee) for employee in employees],
        'selected': _employee_payload(selected) if selected else None,
        'history': [_history_payload(item) for item in history],
        'can_manage': is_hr(request.user),
    })

# API: Leaves - Get Balance and Request Leave
@login_required
def api_leaves(request):
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee profile missing'}, status=404)
        
    if request.method == 'GET':
        # Get balances
        balances = LeaveBalance.objects.filter(employee=employee).select_related('leave_type')
        balances_data = []
        for b in balances:
            balances_data.append({
                'leave_type': b.leave_type.name,
                'remaining_days': b.remaining_days,
                'max_days': b.leave_type.max_days,
            })
            
        # Get history
        requests = LeaveRequest.objects.filter(employee=employee).select_related('leave_type', 'approved_by').order_by('-start_date')
        requests_data = []
        for r in requests:
            requests_data.append({
                'id': r.id,
                'leave_type': r.leave_type.name,
                'start_date': r.start_date.strftime('%Y-%m-%d'),
                'end_date': r.end_date.strftime('%Y-%m-%d'),
                'requested_days': r.requested_days,
                'status': r.status,
                'comments': r.comments,
                'approved_by': r.approved_by.get_full_name() if r.approved_by else None
            })
            
        # Get available leave types for dropdown
        leave_types = LeaveType.objects.all()
        types_data = [{'id': t.id, 'name': t.name, 'max_days': t.max_days} for t in leave_types]
        
        return JsonResponse({
            'success': True,
            'balances': balances_data,
            'requests': requests_data,
            'leave_types': types_data
        })
        
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            leave_type_id = data.get('leave_type_id')
            start_date_str = data.get('start_date')
            end_date_str = data.get('end_date')
            comments = data.get('comments', '')
            
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Invalid form data'}, status=400)
            
        leave_type = get_object_or_404(LeaveType, id=leave_type_id)
        
        # Calculate requested days
        requested_days = (end_date - start_date).days + 1
        if requested_days <= 0:
            return JsonResponse({'success': False, 'error': 'End date must be after or equal to start date'}, status=400)
            
        # Check leave balance
        balance, created = LeaveBalance.objects.get_or_create(
            employee=employee,
            leave_type=leave_type,
            defaults={'remaining_days': leave_type.max_days, 'last_updated': timezone.localdate()}
        )
        
        if balance.remaining_days < requested_days:
            return JsonResponse({'success': False, 'error': f'Insufficient leave balance. You only have {balance.remaining_days} days remaining.'}, status=400)
            
        # Create Leave Request
        req = LeaveRequest.objects.create(
            employee=employee,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            requested_days=requested_days,
            status='Pending',
            comments=comments
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Leave request submitted successfully!',
            'request': {
                'id': req.id,
                'leave_type': req.leave_type.name,
                'start_date': req.start_date.strftime('%Y-%m-%d'),
                'end_date': req.end_date.strftime('%Y-%m-%d'),
                'requested_days': req.requested_days,
                'status': req.status
            }
        })

# API: Admin Leaves - List all pending, and Approve/Reject (HR Only)
@login_required
def api_admin_leaves(request):
    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
    if request.method == 'GET':
        requests = LeaveRequest.objects.filter(status='Pending').select_related('employee', 'leave_type').order_by('start_date')
        data = []
        for r in requests:
            data.append({
                'id': r.id,
                'employee_name': r.employee.get_full_name(),
                'employee_id': r.employee.id,
                'leave_type': r.leave_type.name,
                'start_date': r.start_date.strftime('%Y-%m-%d'),
                'end_date': r.end_date.strftime('%Y-%m-%d'),
                'requested_days': r.requested_days,
                'comments': r.comments,
            })
        return JsonResponse({'success': True, 'requests': data})
        
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            request_id = data.get('request_id')
            action = data.get('action') # 'approve' or 'reject'
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
            
        req = get_object_or_404(LeaveRequest, id=request_id)
        if req.status != 'Pending':
            return JsonResponse({'success': False, 'error': 'Leave request is already processed'}, status=400)
            
        try:
            hr_employee = Employee.objects.get(user=request.user)
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'HR Employee profile missing'}, status=400)
            
        if action == 'approve':
            # Deduct balance
            balance, _ = LeaveBalance.objects.get_or_create(
                employee=req.employee,
                leave_type=req.leave_type,
                defaults={'remaining_days': req.leave_type.max_days, 'last_updated': timezone.localdate()}
            )
            balance.remaining_days -= req.requested_days
            balance.last_updated = timezone.localdate()
            balance.save()
            
            req.status = 'Approved'
            req.approved_by = hr_employee
            req.approved_date = timezone.localdate()
            req.save()
            return JsonResponse({'success': True, 'message': 'Leave request approved successfully!'})
            
        elif action == 'reject':
            req.status = 'Rejected'
            req.approved_by = hr_employee
            req.approved_date = timezone.localdate()
            req.save()
            return JsonResponse({'success': True, 'message': 'Leave request rejected successfully!'})
            
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)

# API: Payroll System (Own / HR all)
@login_required
def api_payroll(request):
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        # Standard admin with no employee profile
        employee = None
        
    is_hr_user = is_hr(request.user)
    
    if request.method == 'GET':
        if is_hr_user:
            # HR sees all payroll
            payroll_records = Payroll.objects.select_related('employee').order_by('-period_start')
        else:
            if not employee:
                return JsonResponse({'success': True, 'records': []})
            # Employee sees own payroll
            payroll_records = Payroll.objects.filter(employee=employee).order_by('-period_start')
            
        data = []
        for p in payroll_records:
            data.append({
                'id': p.id,
                'employee_name': p.employee.get_full_name(),
                'employee_id': p.employee.id,
                'period_start': p.period_start.strftime('%Y-%m-%d'),
                'period_end': p.period_end.strftime('%Y-%m-%d'),
                'gross_salary': str(p.gross_salary),
                'required_days': p.required_days,
                'days_worked': p.days_worked,
                'deduction_amount': str(p.deduction_amount),
                'deductions': str(p.deductions),
                'net_salary': str(p.net_salary),
                'payment_status': p.payment_status
            })
        return JsonResponse({'success': True, 'records': data})
        
    elif request.method == 'POST':
        # HR triggers payroll calculation
        if not is_hr_user:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
            
        try:
            data = json.loads(request.body)
            start_str = data.get('period_start')
            end_str = data.get('period_end')
            
            period_start = datetime.datetime.strptime(start_str, '%Y-%m-%d').date()
            period_end = datetime.datetime.strptime(end_str, '%Y-%m-%d').date()
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid dates'}, status=400)
            
        # Calculate for all active employees
        active_employees = Employee.objects.filter(is_deleted=False)
        calculated = 0
        
        for emp in active_employees:
            # check if payroll record already exists for this employee and period
            if Payroll.objects.filter(employee=emp, period_start=period_start, period_end=period_end).exists():
                continue
                
            payroll_vals = Payroll.calculate_for_employee(emp, period_start, period_end)
            if payroll_vals['gross_salary'] <= 0:
                continue # Skip employees with no salary registered
                
            Payroll.objects.create(
                employee=emp,
                period_start=period_start,
                period_end=period_end,
                gross_salary=payroll_vals['gross_salary'],
                required_days=payroll_vals['required_days'],
                days_worked=payroll_vals['days_worked'],
                deduction_amount=payroll_vals['deduction_amount'],
                deductions=payroll_vals['deductions'],
                net_salary=payroll_vals['net_salary'],
                payment_status='Pending'
            )
            calculated += 1
            
        return JsonResponse({
            'success': True,
            'message': f'Successfully calculated and generated payroll for {calculated} employees!'
        })

# API: Performance Evaluation
@login_required
def api_performance(request):
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        employee = None
        
    is_hr_user = is_hr(request.user)
    
    if request.method == 'GET':
        if is_hr_user:
            evals = PerformanceEvaluation.objects.select_related('employee', 'evaluator').order_by('-evaluation_date')
        else:
            if not employee:
                return JsonResponse({'success': True, 'evaluations': [], 'peers': [], 'kpis': []})
            # Employee sees own reviews or reviews they performed
            evals = PerformanceEvaluation.objects.filter(employee=employee).select_related('employee', 'evaluator').order_by('-evaluation_date')
            
        evals_data = []
        for ev in evals:
            evals_data.append({
                'id': ev.id,
                'employee_name': ev.employee.get_full_name(),
                'employee_id': ev.employee.id,
                'evaluator_name': ev.evaluator.get_full_name(),
                'evaluation_type': ev.get_evaluation_type_display(),
                'evaluation_date': ev.evaluation_date.strftime('%Y-%m-%d'),
                'overall_score': str(ev.overall_score) if ev.overall_score else 'N/A',
                'outcome': ev.get_outcome_display() if ev.outcome else 'N/A',
                'status': ev.status,
                'comments': ev.comments
            })
            
        # Get active KPI categories and indicators
        categories = KPICategory.objects.all()
        kpis_data = []
        for cat in categories:
            indicators = cat.kpiindicator_set.filter(is_active=True)
            kpis_data.append({
                'category_id': cat.id,
                'category_name': cat.name,
                'weight': str(cat.weight),
                'indicators': [{'id': ind.id, 'name': ind.name, 'description': ind.description, 'max_score': ind.max_score} for ind in indicators]
            })
            
        # Get peers (other employees)
        peers = Employee.objects.filter(is_deleted=False)
        if employee:
            peers = peers.exclude(id=employee.id)
        peers_data = [{'id': p.id, 'full_name': p.get_full_name()} for p in peers]
        
        return JsonResponse({
            'success': True,
            'evaluations': evals_data,
            'kpis': kpis_data,
            'peers': peers_data
        })
        
    elif request.method == 'POST':
        if not employee:
            return JsonResponse({'success': False, 'error': 'Must have an Employee profile to evaluate'}, status=400)
            
        try:
            data = json.loads(request.body)
            target_employee_id = data.get('employee_id')
            eval_type = data.get('evaluation_type', 'peer')
            comments = data.get('comments', '')
            scores_raw = data.get('scores', {}) # { kpi_indicator_id: score_value }
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
            
        target_employee = get_object_or_404(Employee, id=target_employee_id)
        
        # Create PerformanceEvaluation
        evaluation = PerformanceEvaluation.objects.create(
            employee=target_employee,
            evaluator=employee,
            evaluation_type=eval_type,
            evaluation_date=timezone.localdate(),
            period_start=timezone.localdate() - datetime.timedelta(days=365), # Mock period
            period_end=timezone.localdate(),
            comments=comments,
            status='Approved'
        )
        
        # Save scores
        for kpi_id, score_val in scores_raw.items():
            indicator = get_object_or_404(KPIIndicator, id=int(kpi_id))
            EvaluationScore.objects.create(
                evaluation=evaluation,
                kpi=indicator,
                score=float(score_val),
                comment="Automated feedback"
            )
            
        # Run calculations
        evaluation.calculate_overall_score()
        evaluation.apply_outcome()
        
        return JsonResponse({
            'success': True,
            'message': 'Performance evaluation submitted successfully!',
            'score': str(evaluation.overall_score),
            'outcome': evaluation.get_outcome_display()
        })

# API: Recruitment & Auto-Screening (HR Only)
@login_required
def api_recruitment(request):
    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
    # Get all active Job Postings
    jobs = JobPosting.objects.select_related('department').all()
    jobs_data = []
    for j in jobs:
        jobs_data.append({
            'id': j.id,
            'title': j.title,
            'department': j.department.name,
            'min_education': j.min_education,
            'min_experience_years': j.min_experience_years,
            'required_skills': j.required_skills,
            'closing_date': j.closing_date.strftime('%Y-%m-%d')
        })
        
    # Get all applications
    apps = Application.objects.select_related('applicant', 'job').order_by('-applied_date')
    apps_data = []
    for a in apps:
        apps_data.append({
            'id': a.id,
            'applicant_name': a.applicant.get_full_name(),
            'applicant_email': a.applicant.email,
            'job_title': a.job.title,
            'status': a.get_status_display(),
            'applied_date': a.applied_date.strftime('%Y-%m-%d'),
            'education_level': a.education_level,
            'experience_years': a.experience_years,
            'skills': a.skills,
            'screening_score': a.screening_score,
            'screening_notes': a.screening_notes
        })
        
    return JsonResponse({
        'success': True,
        'job_postings': jobs_data,
        'applications': apps_data
    })

# API: System Settings (HR Only)
@login_required
def api_settings(request):
    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
    if request.method == 'GET':
        settings = SystemSetting.objects.all()
        data = []
        for s in settings:
            data.append({
                'id': s.id,
                'category': s.category,
                'key': s.key,
                'value': s.value,
                'description': s.description
            })
        return JsonResponse({'success': True, 'settings': data})
        
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            setting_id = data.get('id')
            value = data.get('value')
        except Exception:
            return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
            
        setting = get_object_or_404(SystemSetting, id=setting_id)
        setting.value = str(value)
        setting.save()
        return JsonResponse({'success': True, 'message': 'Setting updated successfully!'})

# API: Simulate Scan for testing (HR Only)
@login_required
def api_simulate_scan(request):
    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request body'}, status=400)
        
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Generate rotating token
    from hr.utils import generate_daily_token
    token = generate_daily_token(employee)
    
    now = timezone.now()
    now_local = timezone.localtime(now)
    today = timezone.localdate()
    
    from hr.models import Attendance, QRScanLog
    attendance, created = Attendance.objects.get_or_create(
        employee=employee,
        date=today,
        defaults={'time_in': now, 'status': 'Present'},
    )
    
    if created:
        attendance.calculate_status()
        attendance.save()
        QRScanLog.objects.create(
            employee=employee, scan_type='time_in', is_successful=True, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'time_in',
                'employee': employee.get_full_name(),
                'time': now_local.strftime('%H:%M:%S'),
                'message': f"Welcome, {employee.first_name}! Clocked in successfully."
            }
        })
    elif attendance.time_out is None:
        attendance.time_out = now
        attendance.save()
        hours = round((now - attendance.time_in).total_seconds() / 3600, 2)
        QRScanLog.objects.create(
            employee=employee, scan_type='time_out', is_successful=True, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'time_out',
                'employee': employee.get_full_name(),
                'time': now_local.strftime('%H:%M:%S'),
                'hours_worked': hours,
                'message': f"Goodbye, {employee.first_name}! You worked {hours} hours."
            }
        })
    else:
        QRScanLog.objects.create(
            employee=employee, scan_type='duplicate', is_successful=False, ip_address='127.0.0.1'
        )
        return JsonResponse({
            'success': True,
            'data': {
                'result': 'already_complete',
                'employee': employee.get_full_name(),
                'message': 'Attendance already fully recorded for today.'
            }
        })


# ============================================================
#  REFERENCE DATA API (for form dropdowns)
# ============================================================

@login_required
def api_reference_data(request):
    """
    GET /api/reference-data/
    Returns departments, positions, roles, and statuses for HR forms.
    """
    from hr.models import Department, Position, Role, EmployeeStatus
    departments = list(Department.objects.filter(is_deleted=False).values('id', 'name').order_by('name'))
    positions   = list(Position.objects.filter(is_deleted=False).values('id', 'title').order_by('title'))
    roles       = list(Role.objects.filter(is_deleted=False).values('id', 'name').order_by('name'))
    statuses    = list(EmployeeStatus.objects.filter(is_deleted=False).values('id', 'name').order_by('name'))
    return JsonResponse({
        'success': True,
        'departments': departments,
        'positions': positions,
        'roles': roles,
        'statuses': statuses,
    })


# ============================================================
#  SINGLE EMPLOYEE CREATE (HR DASHBOARD FORM)
# ============================================================

@login_required
def api_create_employee_hr(request):
    """
    POST /api/employees/create-hr/
    Creates a single employee from the HR dashboard form.
    Performs full field-level validation and sends a welcome email.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON body'}, status=400)

    # ── Field-level validation ──────────────────────────────
    from django.core.exceptions import ValidationError as DjangoValidationError
    from hr.services.employee_service import create_employee as svc_create_employee

    try:
        employee = svc_create_employee(data)
    except DjangoValidationError as e:
        payload = {'success': False, 'error': 'Validation failed.'}
        if hasattr(e, 'message_dict'):
            payload['errors'] = {field: messages[0] for field, messages in e.message_dict.items()}
        else:
            payload['error'] = str(e.message if hasattr(e, 'message') else e)
        return JsonResponse(payload, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to create employee: {str(e)}'}, status=500)

    send_email = bool(data.get('send_email', True))
    email = (data.get('email') or '').strip().lower()
    return JsonResponse({
        'success': True,
        'message': f"Employee {employee.get_full_name()} created successfully! Login credentials have been emailed." if send_email else f"Employee {employee.get_full_name()} created successfully!",
        'employee': {
            'id': employee.id,
            'employee_id': employee.employee_id,
            'full_name': employee.get_full_name(),
            'email': employee.user.email if employee.user else email,
            'username': employee.user.username if employee.user else '',
            'pin': employee.pin or employee.attendance_pin,
            'department': employee.department.name if employee.department else 'N/A',
            'position': employee.position.title if employee.position else 'N/A',
        }
    }, status=201)

    errors = {}
    first_name = (data.get('first_name') or '').strip()
    last_name  = (data.get('last_name')  or '').strip()
    email      = (data.get('email')      or '').strip().lower()
    username   = (data.get('username')   or '').strip()
    phone      = (data.get('phone')      or '').strip()
    address    = (data.get('address')    or '').strip()
    hire_date  = (data.get('hire_date')  or '').strip() or None
    dept_id    = data.get('department_id')
    pos_id     = data.get('position_id')
    role_id    = data.get('role_id')
    status_id  = data.get('status_id')
    send_email = bool(data.get('send_email', True))
    signature_data = (data.get('signature_data') or '').strip()

    if not first_name:
        errors['first_name'] = 'First name is required.'
    if not last_name:
        errors['last_name'] = 'Last name is required.'

    import re
    if not email:
        errors['email'] = 'Email address is required.'
    elif not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        errors['email'] = 'Enter a valid email address.'
    elif User.objects.filter(email__iexact=email).exists():
        errors['email'] = 'A user with this email already exists.'

    if username and User.objects.filter(username__iexact=username).exists():
        errors['username'] = f"Username '{username}' is already taken."

    if signature_data and not signature_data.startswith('data:image/'):
        errors['signature_data'] = 'Invalid signature format.'

    if errors:
        return JsonResponse({'success': False, 'errors': errors, 'error': 'Validation failed.'}, status=400)

    # ── Build data dict for service layer ──────────────────
    from hr.services.employee_service import create_employee as svc_create_employee
    from django.core.exceptions import ValidationError as DjangoValidationError

    service_data = {
        'first_name':    first_name,
        'last_name':     last_name,
        'email':         email,
        'username':      username or None,
        'phone':         phone,
        'address':       address,
        'hire_date':     hire_date,
        'department_id': dept_id,
        'position_id':   pos_id,
        'role_id':       role_id,
        'status_id':     status_id,
        'signature_data': signature_data,
        'send_email':     send_email,
    }

    try:
        employee = svc_create_employee(service_data)
    except DjangoValidationError as e:
        return JsonResponse({'success': False, 'error': str(e.message if hasattr(e, 'message') else e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to create employee: {str(e)}'}, status=500)

    return JsonResponse({
        'success': True,
        'message': f"Employee {employee.get_full_name()} created successfully! Login credentials have been emailed." if send_email else f"Employee {employee.get_full_name()} created successfully!",
        'employee': {
            'id': employee.id,
            'employee_id': employee.employee_id,
            'full_name': employee.get_full_name(),
            'email': employee.user.email if employee.user else email,
            'username': employee.user.username if employee.user else '',
            'pin': employee.pin or employee.attendance_pin,
            'department': employee.department.name if employee.department else 'N/A',
            'position': employee.position.title if employee.position else 'N/A',
        }
    }, status=201)


# ============================================================
#  BULK ONBOARDING & TABLET SIGNATURE ATTENDANCE
# ============================================================

from django.views.decorators.csrf import csrf_exempt
import csv
import io
import base64
import uuid
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.crypto import get_random_string

@login_required
def api_upload_employees_csv(request):
    """
    POST /api/employees/upload-csv/
    Allows HR managers to upload a CSV (.csv) or Excel (.xlsx) spreadsheet
    to bulk-create employee profiles.

    Required columns: username, email, first_name, last_name
    Optional columns: phone, department, position, role, hire_date
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    if not is_hr(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    uploaded_file = request.FILES.get('csv_file')
    if not uploaded_file:
        return JsonResponse({'success': False, 'error': 'No file uploaded.'}, status=400)

    from django.core.exceptions import ValidationError as DjangoValidationError
    from hr.services.employee_service import bulk_create_employees_from_file

    try:
        result = bulk_create_employees_from_file(uploaded_file)
    except DjangoValidationError as e:
        return JsonResponse({'success': False, 'error': str(e.message if hasattr(e, 'message') else e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to upload roster: {str(e)}'}, status=500)

    return JsonResponse({
        'success': True,
        'message': f"Successfully onboarded {result['created_count']} employee(s).",
        'created': result['created'],
        'errors': result['errors']
    })

    file_name = uploaded_file.name.lower()
    if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Invalid file format. Please upload a .csv or .xlsx spreadsheet file.'
        }, status=400)

    rows = []
    required_headers = ['username', 'email', 'first_name', 'last_name']

    # ── Parse CSV ──────────────────────────────────────────
    if file_name.endswith('.csv'):
        try:
            csv_data = uploaded_file.read().decode('utf-8-sig')
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to decode CSV: {str(e)}'}, status=400)
        reader = csv.DictReader(io.StringIO(csv_data))
        missing = [h for h in required_headers if h not in (reader.fieldnames or [])]
        if missing:
            return JsonResponse({
                'success': False,
                'error': f"Missing required column headers: {', '.join(missing)}",
            }, status=400)
        rows = list(reader)

    # ── Parse Excel (.xlsx / .xls) ─────────────────────────
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            missing = [h for h in required_headers if h not in headers]
            if missing:
                return JsonResponse({
                    'success': False,
                    'error': f"Missing required column headers: {', '.join(missing)}",
                }, status=400)
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}
                if any(row_dict.values()):  # skip fully empty rows
                    rows.append(row_dict)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to parse spreadsheet: {str(e)}'}, status=400)

    # ── Try to get the current HR employee for history ─────
    try:
        hr_employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        hr_employee = None

    success_count = 0
    errors = []
    created_employees = []

    import random
    for idx, row in enumerate(rows, start=2):
        username   = (row.get('username')   or '').strip()
        email      = (row.get('email')      or '').strip()
        first_name = (row.get('first_name') or '').strip()
        last_name  = (row.get('last_name')  or '').strip()

        if not username:
            errors.append(f"Row {idx}: Username is required.")
            continue
        if not email:
            errors.append(f"Row {idx} ({username}): Email is required.")
            continue

        if User.objects.filter(username=username).exists():
            errors.append(f"Row {idx} ({username}): Username '{username}' already exists — skipped.")
            continue
        if User.objects.filter(email=email).exists():
            errors.append(f"Row {idx} ({username}): Email '{email}' already exists — skipped.")
            continue

        try:
            with transaction.atomic():
                password = get_random_string(12)
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=password
                )
                emp = Employee.objects.get(user=user)
                emp.first_name = first_name
                emp.last_name  = last_name

                # Unique employee ID
                emp_id_candidate = f"EMP-{random.randint(100000, 999999)}"
                while Employee.objects.filter(employee_id=emp_id_candidate).exists():
                    emp_id_candidate = f"EMP-{random.randint(100000, 999999)}"
                emp.employee_id = emp_id_candidate

                # Auto PIN
                pin = f"{random.randint(0, 999999):06d}"
                emp.pin = pin
                emp.attendance_pin = pin

                # Optional phone/hire_date from sheet
                if row.get('phone'):
                    emp.phone = row['phone'].strip()
                if row.get('hire_date'):
                    try:
                        import datetime
                        emp.hire_date = datetime.datetime.strptime(row['hire_date'].strip(), '%Y-%m-%d').date()
                    except Exception:
                        pass
                emp.save()

                EmployeeHistory.objects.create(
                    employee=emp,
                    event_type='hired',
                    start_date=timezone.localdate(),
                    new_value='Joined via bulk upload.',
                    notes=f"Auto-generated PIN: {pin}",
                    recorded_by=hr_employee
                )

                # Send welcome email
                from hr.services.employee_service import send_employee_credentials
                send_employee_credentials(emp, email, pin, password=password)

                success_count += 1
                created_employees.append({
                    'employee_id': emp_id_candidate,
                    'name': emp.get_full_name(),
                    'email': email,
                    'pin': pin,
                })
        except Exception as e:
            errors.append(f"Row {idx} ({username}): Error — {str(e)}")

    return JsonResponse({
        'success': True,
        'message': f"Successfully onboarded {success_count} employee(s).",
        'created': created_employees,
        'errors': errors
    })


@csrf_exempt
def api_tablet_authenticate(request):
    """
    POST /api/attendance/tablet/authenticate/
    Validates employee ID and PIN for tablet kiosk, returns profile details.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        employee_id = (data.get('employee_id') or '').strip()
        pin = (data.get('pin') or '').strip()
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request body.'}, status=400)
        
    if not employee_id or not pin:
        return JsonResponse({'success': False, 'error': 'Employee ID and PIN are required.'}, status=400)
        
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid Employee ID or PIN.'}, status=404)
        
    if (employee.attendance_pin or employee.pin) != pin:
        return JsonResponse({'success': False, 'error': 'Invalid Employee ID or PIN.'}, status=401)
        
    photo_url = employee.photo.url if employee.photo else 'https://images.unsplash.com/photo-1535713875002-d1d0cf377fde?auto=format&fit=crop&w=150&h=150'
    
    return JsonResponse({
        'success': True,
        'employee': {
            'id': employee.id,
            'employee_id': employee.employee_id,
            'full_name': employee.get_full_name(),
            'photo_url': photo_url,
            'department': employee.department.name if employee.department else 'General',
            'position': employee.position.title if employee.position else 'Staff'
        }
    })


@csrf_exempt
def api_tablet_submit(request):
    """
    POST /api/attendance/tablet/submit/
    Saves tablet signature canvas base64 image and logs clock-in/clock-out events.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        employee_id = (data.get('employee_id') or '').strip()
        pin = (data.get('pin') or '').strip()
        signature_data = data.get('signature')
        action = data.get('action') # 'check_in' or 'check_out'
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request body.'}, status=400)
        
    if not employee_id or not pin or not signature_data or not action:
        return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)
        
    from hr.utils import verify_employee_credentials
    employee, verification_log, verification_error = verify_employee_credentials(
        employee_id,
        pin,
        signature_data,
    )
    if verification_error:
        from hr.models import QRScanLog
        if employee:
            QRScanLog.objects.create(
                employee=employee,
                scan_type='tablet_verification_failed',
                is_successful=False,
                ip_address=request.META.get('REMOTE_ADDR', ''),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                failure_reason=verification_error,
            )
        return JsonResponse({'success': False, 'error': verification_error, 'verification': verification_log}, status=401)
        
    # Decode base64 signature
    try:
        format, imgstr = signature_data.split(';base64,')
        ext = format.split('/')[-1]
        decoded_file = base64.b64decode(imgstr)
        file_name = f"sig_{employee.employee_id}_{uuid.uuid4().hex[:8]}.{ext}"
        signature_file = ContentFile(decoded_file, name=file_name)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to process signature canvas: {str(e)}'}, status=400)
        
    from hr.models import Attendance, QRScanLog
    today = timezone.localdate()
    now = timezone.now()
    now_local = timezone.localtime(now)
    
    attendance = Attendance.objects.filter(employee=employee, date=today).first()
    
    if action == 'check_in':
        if attendance:
            return JsonResponse({'success': False, 'error': 'Already checked in for today.'}, status=400)
            
        attendance = Attendance.objects.create(
            employee=employee,
            date=today,
            time_in=now,
            signature=signature_file,
            status='Present'
        )
        attendance.calculate_status()
        attendance.save()
        
        QRScanLog.objects.create(
            employee=employee,
            scan_type='time_in',
            is_successful=True,
            ip_address=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            failure_reason=f"Tablet check-in signature upload; verification={verification_log}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f"Welcome, {employee.first_name}! Clocked in successfully at {now_local.strftime('%H:%M:%S')}."
        })
        
    elif action == 'check_out':
        if not attendance:
            return JsonResponse({'success': False, 'error': 'No check-in record found for today. You must check in first.'}, status=400)
            
        if attendance.time_out is not None:
            return JsonResponse({'success': False, 'error': 'Already clocked out for today.'}, status=400)
            
        attendance.time_out = now
        if signature_file:
            attendance.signature = signature_file
        attendance.save()
        
        hours = round((now - attendance.time_in).total_seconds() / 3600, 2)
        
        QRScanLog.objects.create(
            employee=employee,
            scan_type='time_out',
            is_successful=True,
            ip_address=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            failure_reason=f"Tablet check-out signature upload; verification={verification_log}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f"Goodbye, {employee.first_name}! Clocked out successfully at {now_local.strftime('%H:%M:%S')}. Worked {hours} hours today."
        })
    else:
        return JsonResponse({'success': False, 'error': 'Invalid attendance action.'}, status=400)


def tablet_kiosk_view(request):
    """
    Renders the tablet-friendly PIN + Signature attendance kiosk view.
    """
    return render(request, 'hr/tablet_kiosk.html')
