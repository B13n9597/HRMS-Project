# hr/services/employee_service.py

import csv
import io
import random
import re

from django.conf import settings
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils import timezone

from hr.models import (
    Attendance,
    Department,
    Employee,
    EmployeeHistory,
    EmployeeStatus as EmployeeStatusModel,
    PerformanceEvaluation,
    Position,
    Role,
    Salary,
)


class EmployeeStatus:
    ACTIVE = "active"
    ON_LEAVE = "on_leave"
    RETIRED = "retired"
    TERMINATED = "terminated"


VALID_TRANSITIONS = {
    EmployeeStatus.ACTIVE: [EmployeeStatus.ON_LEAVE, EmployeeStatus.RETIRED, EmployeeStatus.TERMINATED],
    EmployeeStatus.ON_LEAVE: [EmployeeStatus.ACTIVE, EmployeeStatus.TERMINATED],
    EmployeeStatus.RETIRED: [],
    EmployeeStatus.TERMINATED: [EmployeeStatus.ACTIVE],
}

STATUS_LABELS = {
    EmployeeStatus.ACTIVE: "Active",
    EmployeeStatus.ON_LEAVE: "On Leave",
    EmployeeStatus.RETIRED: "Retired",
    EmployeeStatus.TERMINATED: "Terminated",
}

ROLE_EMPLOYEE = "employee"
ROLE_HR = "hr"
ROLE_DEAN = "dean"

ROLE_REDIRECTS = {
    ROLE_EMPLOYEE: "/dashboard/employee/",
    ROLE_HR: "/dashboard/hr/",
    ROLE_DEAN: "/dashboard/dean/",
}


@transaction.atomic
def create_employee(data: dict) -> Employee:
    validate_employee_payload(data)

    first_name = data["first_name"].strip()
    last_name = data["last_name"].strip()
    email = data["email"].lower().strip()
    attendance_pin = str(data.get("attendance_pin") or data.get("pin") or _generate_pin()).strip()
    signature_data = (data.get("signature_data") or data.get("signature") or "").strip()
    phone = (data.get("phone") or "").strip()
    send_email_flag = bool(data.get("send_email", True))

    if not attendance_pin.isdigit() or len(attendance_pin) != 6:
        raise ValidationError("attendance_pin must be exactly 6 numeric digits.")
    if signature_data and not signature_data.startswith("data:image/"):
        raise ValidationError("signature_data must be a Base64 image data URL.")
    if User.objects.filter(email__iexact=email).exists():
        raise ValidationError("A user with this email already exists.")

    # Optional phone validation (basic international-ish pattern)
    if phone:
        phone_pattern = re.compile(r"^\+?[0-9 \-()]{6,25}$")
        if not phone_pattern.match(phone):
            raise ValidationError("Phone number format is invalid.")

    username = _unique_username(data.get("username") or email.split("@")[0])
    password = data.get("password") or get_random_string(12)
    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
    )

    employee_id = (data.get("employee_id") or "").strip() or _generate_employee_id()
    employee, _ = Employee.objects.get_or_create(
        user=user,
        defaults={
            'first_name': first_name,
            'last_name': last_name,
            'employee_id': employee_id,
            'pin': attendance_pin,
            'attendance_pin': attendance_pin,
            'signature_data': signature_data,
            'department': _resolve_fk(Department, data.get("department") or data.get("department_id")),
            'position': _resolve_fk(Position, data.get("position") or data.get("position_id")),
            'role': _resolve_fk(Role, data.get("role") or data.get("role_id")),
            'status': _resolve_status(data.get("status") or data.get("status_id")),
            'phone': (data.get("phone") or "").strip(),
            'address': (data.get("address") or "").strip(),
            'hire_date': data.get("hire_date") or data.get("date_joined"),
        }
    )

    employee.first_name = first_name
    employee.last_name = last_name
    employee.employee_id = employee_id
    employee.pin = attendance_pin
    employee.attendance_pin = attendance_pin
    employee.signature_data = signature_data
    employee.department = _resolve_fk(Department, data.get("department") or data.get("department_id"))
    employee.position = _resolve_fk(Position, data.get("position") or data.get("position_id"))
    employee.role = _resolve_fk(Role, data.get("role") or data.get("role_id"))
    employee.status = _resolve_status(data.get("status") or data.get("status_id"))
    employee.phone = (data.get("phone") or "").strip()
    employee.address = (data.get("address") or "").strip()

    if data.get("hire_date") or data.get("date_joined"):
        employee.hire_date = data.get("hire_date") or data.get("date_joined")

    employee.save()

    # ── Optionally create an initial Salary record ──────────────
    raw_salary = data.get("base_salary") or data.get("initial_salary")
    if raw_salary:
        try:
            base_salary_val = float(str(raw_salary).replace(',', ''))
            if base_salary_val > 0:
                Salary.objects.create(
                    employee=employee,
                    base_salary=base_salary_val,
                    effective_from=employee.hire_date or timezone.localdate(),
                )
        except (ValueError, TypeError):
            pass  # Skip invalid salary values silently

    # Send credentials email only when requested
    if send_email_flag:
        send_employee_credentials(employee, email, attendance_pin, password=password)
    return employee


def validate_employee_payload(data: dict) -> None:
    errors = {}
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    email = (data.get("email") or "").strip().lower()

    if not first_name:
        errors["first_name"] = "First name is required."
    if not last_name:
        errors["last_name"] = "Last name is required."
    if not email:
        errors["email"] = "Email is required."
    elif not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        errors["email"] = "Enter a valid email address."
    elif User.objects.filter(email__iexact=email).exists():
        errors["email"] = "A user with this email already exists."

    username = (data.get("username") or "").strip()
    if username and User.objects.filter(username__iexact=username).exists():
        errors["username"] = f"Username '{username}' is already taken."

    if not (data.get("department") or data.get("department_id")):
        errors["department_id"] = "Department is required."
    if not (data.get("role") or data.get("role_id")):
        errors["role_id"] = "Role is required."

    if errors:
        raise ValidationError(errors)


@transaction.atomic
def bulk_create_employees_from_file(uploaded_file) -> dict:
    rows = _read_employee_upload_rows(uploaded_file)
    created = []
    errors = []

    for row_number, row in rows:
        try:
            employee = create_employee(_employee_data_from_upload_row(row))
            record_lifecycle_event(
                employee,
                "hired",
                "Joined via bulk upload.",
            )
            created.append({
                "employee_id": employee.employee_id,
                "name": employee.get_full_name(),
                "email": employee.user.email if employee.user else "",
                "pin": employee.pin or employee.attendance_pin,
            })
        except ValidationError as exc:
            errors.append(f"Row {row_number}: {_validation_error_text(exc)}")
        except Exception as exc:
            errors.append(f"Row {row_number}: {exc}")

    return {
        "created_count": len(created),
        "created": created,
        "errors": errors,
    }


@transaction.atomic
def update_employee(employee_id: int, data: dict) -> Employee:
    employee = get_employee_by_id(employee_id)
    _apply_employee_data(employee, data)
    employee.save()

    if employee.user:
        employee.user.first_name = employee.first_name
        employee.user.last_name = employee.last_name
        if data.get("email"):
            employee.user.email = data["email"].lower().strip()
        if data.get("password"):
            employee.user.set_password(data["password"])
        employee.user.save()

    return employee


def update_employee_details(employee_id: int, data: dict) -> Employee:
    return update_employee(employee_id, data)


@transaction.atomic
def delete_employee(employee_id: int) -> None:
    employee = get_employee_by_id(employee_id)
    if employee.user:
        employee.user.is_active = False
        employee.user.save(update_fields=["is_active"])
    employee.delete()


@transaction.atomic
def transition_employee_status(employee_id: int, new_status: str) -> Employee:
    employee = get_employee_by_id(employee_id)
    normalized_status = _normalize_status_key(new_status)
    current_status = _employee_status_key(employee)
    allowed = VALID_TRANSITIONS.get(current_status, [])

    if normalized_status not in allowed and normalized_status != current_status:
        raise ValidationError(f"Invalid transition: {current_status} to {normalized_status}")

    status_obj, _ = EmployeeStatusModel.objects.get_or_create(name=STATUS_LABELS[normalized_status])
    employee.status = status_obj
    employee.save(update_fields=["status"])
    return employee


def activate_employee(employee_id: int) -> Employee:
    return transition_employee_status(employee_id, EmployeeStatus.ACTIVE)


def put_employee_on_leave(employee_id: int) -> Employee:
    return transition_employee_status(employee_id, EmployeeStatus.ON_LEAVE)


def terminate_employee(employee_id: int) -> Employee:
    return transition_employee_status(employee_id, EmployeeStatus.TERMINATED)


def record_lifecycle_event(employee: Employee, event_type: str, notes: str = "", recorded_by: Employee = None):
    if event_type not in dict(EmployeeHistory.EVENT_CHOICES):
        return None
    return EmployeeHistory.objects.create(
        employee=employee,
        department=employee.department,
        position=employee.position,
        event_type=event_type,
        new_value=employee.status.name if employee.status else "",
        notes=notes,
        recorded_by=recorded_by,
        start_date=timezone.localdate(),
    )


def get_employee_by_id(employee_id: int) -> Employee:
    try:
        return Employee.objects.select_related("user", "role", "department", "position", "status").get(id=employee_id)
    except Employee.DoesNotExist:
        raise ValidationError(f"Employee {employee_id} not found.")


def get_employee(employee_id: int) -> Employee:
    """Returns a single employee by primary key."""
    return get_employee_by_id(employee_id)


def set_on_leave(employee_id: int) -> Employee:
    """Sets employee status to On Leave."""
    return put_employee_on_leave(employee_id)


def get_employee_for_user(user):
    try:
        return Employee.objects.select_related("user", "role", "department", "position", "status").get(user=user)
    except Employee.DoesNotExist:
        return None


def get_all_employees():
    return Employee.objects.select_related("user", "role", "department", "position", "status").order_by(
        "last_name", "first_name", "id"
    )


def get_all_active_employees():
    return get_all_employees().filter(status__name__iexact="Active")


def get_all_attendance_logs():
    return Attendance.objects.select_related("employee", "employee__department").order_by("-date", "-time_in")


def get_reference_data():
    return {
        "departments": Department.objects.order_by("name"),
        "positions": Position.objects.order_by("title"),
        "roles": Role.objects.order_by("name"),
        "statuses": EmployeeStatusModel.objects.order_by("name"),
    }


def get_dean_report_summary():
    employees = get_all_employees()
    evaluations = PerformanceEvaluation.objects.select_related("employee", "evaluator").order_by("-evaluation_date")[:25]
    return {
        "employees": employees,
        "employee_count": employees.count(),
        "active_count": employees.filter(status__name__iexact="Active").count(),
        "on_leave_count": employees.filter(status__name__iexact="On Leave").count(),
        "terminated_count": employees.filter(status__name__iexact="Terminated").count(),
        "performance_evaluations": evaluations,
    }


def get_role_key(user) -> str:
    if user.is_superuser:
        return ROLE_HR

    employee = get_employee_for_user(user)
    role_name = employee.role.name.strip().lower() if employee and employee.role else ""

    if role_name in {"dean", "academic dean"}:
        return ROLE_DEAN
    if role_name in {"hr", "admin", "hr manager", "human resources"}:
       return ROLE_HR
    return ROLE_EMPLOYEE


def get_dashboard_redirect(user) -> str:
    return ROLE_REDIRECTS[get_role_key(user)]


def can_manage_employees(user) -> bool:
    return get_role_key(user) == ROLE_HR


def can_view_dean_reports(user) -> bool:
    return get_role_key(user) in {ROLE_DEAN, ROLE_HR}


def send_employee_credentials(employee: Employee, recipient_email: str, attendance_pin: str, password: str = "") -> None:
    send_mail(
        subject="Your ACT HRMS credentials",
        message=(
            f"Hello {employee.get_full_name()},\n\n"
            "Your ACT HRMS profile has been created.\n\n"
            f"Username: {employee.user.username if employee.user else ''}\n"
            f"Temporary password: {password}\n"
            f"Employee ID: {employee.employee_id}\n"
            f"Attendance PIN: {attendance_pin}\n\n"
            "Use your username and password to sign in to HRMS."
        ),
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        recipient_list=[recipient_email],
        fail_silently=getattr(settings, "EMAIL_FAIL_SILENTLY", True),
    )


def _apply_employee_data(employee: Employee, data: dict) -> None:
    allowed = ("first_name", "last_name", "phone", "address", "employee_id", "attendance_pin", "pin")
    updated = False

    for field in allowed:
        if field in data:
            value = data.get(field)
            setattr(employee, field, value.strip() if isinstance(value, str) else value)
            updated = True

    if data.get("department") or data.get("department_id"):
        employee.department = _resolve_fk(Department, data.get("department") or data.get("department_id"))
        updated = True
    if data.get("position") or data.get("position_id"):
        employee.position = _resolve_fk(Position, data.get("position") or data.get("position_id"))
        updated = True
    if data.get("role") or data.get("role_id"):
        employee.role = _resolve_fk(Role, data.get("role") or data.get("role_id"))
        updated = True
    if data.get("status") or data.get("status_id"):
        employee.status = _resolve_status(data.get("status") or data.get("status_id"))
        updated = True
    if data.get("hire_date"):
        employee.hire_date = data.get("hire_date")
        updated = True

    if not updated and not data.get("email") and not data.get("password"):
        raise ValidationError("No valid fields provided for update.")


def _generate_pin() -> str:
    return f"{random.SystemRandom().randint(0, 999999):06d}"


def _generate_employee_id() -> str:
    while True:
        candidate = f"EMP-{random.SystemRandom().randint(1000, 9999)}"
        if not Employee.objects.filter(employee_id=candidate).exists():
            return candidate


def _unique_username(base_username: str) -> str:
    base = "".join(ch for ch in base_username.lower().strip() if ch.isalnum() or ch in "._-") or "employee"
    username = base
    suffix = 1
    while User.objects.filter(username=username).exists():
        suffix += 1
        username = f"{base}{suffix}"
    return username


def _resolve_fk(model, value):
    if not value:
        return None
    if isinstance(value, model):
        return value
    try:
        return model.objects.get(id=value)
    except (model.DoesNotExist, TypeError, ValueError):
        raise ValidationError(f"Invalid {model.__name__} id: {value}")


def _resolve_status(value):
    if not value:
        status, _ = EmployeeStatusModel.objects.get_or_create(name="Active")
        return status
    if isinstance(value, EmployeeStatusModel):
        return value
    try:
        if str(value).isdigit():
            return EmployeeStatusModel.objects.get(id=value)
        return EmployeeStatusModel.objects.get(name__iexact=str(value))
    except EmployeeStatusModel.DoesNotExist:
        raise ValidationError(f"Invalid EmployeeStatus: {value}")


def _employee_status_key(employee: Employee) -> str:
    if not employee.status:
        return EmployeeStatus.ACTIVE
    return _normalize_status_key(employee.status.name)


def _normalize_status_key(value: str) -> str:
    normalized = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "active": EmployeeStatus.ACTIVE,
        "on_leave": EmployeeStatus.ON_LEAVE,
        "leave": EmployeeStatus.ON_LEAVE,
        "retired": EmployeeStatus.RETIRED,
        "terminated": EmployeeStatus.TERMINATED,
    }
    if normalized not in aliases:
        raise ValidationError(f"Invalid status: {value}")
    return aliases[normalized]


def _read_employee_upload_rows(uploaded_file):
    if not uploaded_file:
        raise ValidationError("No file uploaded.")

    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        try:
            text = uploaded_file.read().decode("utf-8-sig")
        except Exception as exc:
            raise ValidationError(f"Failed to decode CSV: {exc}")

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValidationError("Uploaded CSV has no header row.")
        return [(index, row) for index, row in enumerate(reader, start=2) if any((row or {}).values())]

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError:
            raise ValidationError("Excel upload requires openpyxl. Install openpyxl on the machine running this project.")

        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            raise ValidationError("Uploaded spreadsheet has no header row.")

        headers = [str(cell.value or "").strip().lower() for cell in header_row]
        rows = []
        for index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {headers[i]: (str(value).strip() if value is not None else "") for i, value in enumerate(values)}
            if any(row.values()):
                rows.append((index, row))
        return rows

    raise ValidationError("Invalid file format. Please upload a .csv or .xlsx spreadsheet file.")


def _employee_data_from_upload_row(row: dict) -> dict:
    full_name = (row.get("name") or row.get("full_name") or "").strip()
    first_name = (row.get("first_name") or "").strip()
    last_name = (row.get("last_name") or "").strip()

    if full_name and not first_name:
        parts = full_name.split()
        first_name = parts[0]
        last_name = " ".join(parts[1:]) or "-"

    department = _resolve_by_name_or_id(Department, row.get("department_id") or row.get("department"))
    role = _resolve_by_name_or_id(Role, row.get("role_id") or row.get("role"))
    position = _resolve_by_name_or_id(Position, row.get("position_id") or row.get("position"))
    status = _resolve_by_name_or_id(EmployeeStatusModel, row.get("status_id") or row.get("status"))

    return {
        "first_name": first_name,
        "last_name": last_name,
        "email": row.get("email"),
        "username": row.get("username"),
        "password": row.get("password"),
        "phone": row.get("phone"),
        "address": row.get("address"),
        "hire_date": row.get("hire_date"),
        "employee_id": row.get("employee_id"),
        "attendance_pin": row.get("attendance_pin") or row.get("pin"),
        "department_id": department.id if department else "",
        "role_id": role.id if role else "",
        "position_id": position.id if position else "",
        "status_id": status.id if status else "",
        "send_email": row.get("send_email", True),
        # Pass base_salary / initial_salary from spreadsheet column if present
        "base_salary": row.get("base_salary") or row.get("salary") or row.get("initial_salary"),
    }


def _resolve_by_name_or_id(model, value):
    value = str(value or "").strip()
    if not value:
        return None
    if value.isdigit():
        return _resolve_fk(model, value)

    lookup = "title__iexact" if model is Position else "name__iexact"
    try:
        return model.objects.get(**{lookup: value})
    except model.DoesNotExist:
        raise ValidationError(f"Invalid {model.__name__}: {value}")


def _validation_error_text(exc):
    if hasattr(exc, "message_dict"):
        return "; ".join(
            f"{field}: {', '.join(messages)}"
            for field, messages in exc.message_dict.items()
        )
    if hasattr(exc, "messages"):
        return "; ".join(exc.messages)
    return str(exc)


def promote_employee(employee_id, position_id, recorder=None):

    employee = Employee.objects.get(id=employee_id)

    old_position = employee.position.title if employee.position else "None"

    new_position = Position.objects.get(id=position_id)

    employee.position = new_position
    employee.save()

    record_lifecycle_event(
        employee,
        "promoted",
        f"Promoted from {old_position} to {new_position.title}",
        recorder
    )

    return employee


def transfer_employee(employee_id, department_id, recorder=None):

    employee = Employee.objects.get(id=employee_id)

    old_department = employee.department.name if employee.department else "None"

    new_department = Department.objects.get(id=department_id)

    employee.department = new_department
    employee.save()

    record_lifecycle_event(
        employee,
        "transferred",
        f"Transferred from {old_department} to {new_department.name}",
        recorder
    )

    return employee
